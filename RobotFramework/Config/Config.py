import pandas as pd
import xlrd
import os
import subprocess
import sys
from pathlib import Path
import datetime
from sqlalchemy import create_engine, types
import sqlite3
import shutil
from datetime import date
from shutil import copyfile
strProjectWorkingDirectory = str(Path(__file__).parents[1])
sys.path.append(strProjectWorkingDirectory)
from Utilities.Email import SendEmail, Generate_Notification_Email
from Utilities.Jira_InitialSetup_config import JiraInitialSetup
from Utilities.helpers import decrypt_data
from openpyxl import load_workbook
import openpyxl
import xlsxwriter

class RunManager():
    def ReadConfig():
        # Read Config sheets
        strProjectWorkingDirectory = str(Path(__file__).parents[1])
        strRunManagerFilePath = strProjectWorkingDirectory + "\Config\RunManager.xlsx"
        # strProjectWorkingDirectory = os.path.abspath(os.getcwd())
        # strRunManagerFilePath = strProjectWorkingDirectory + "\Config\RunManager.xlsx"
        wb = xlrd.open_workbook(strRunManagerFilePath)
        sheet = wb.sheet_by_name("Config")

        # read all the config parameters
        strModule = sheet.cell_value(6, 3)
        strRegion = sheet.cell_value(6, 7)
        os.environ['Region'] = strRegion
        strEnvironment = sheet.cell_value(8, 3)
        os.environ['Env'] = strEnvironment
        strJiraProject = sheet.cell_value(8, 7)
        strVersionName = sheet.cell_value(10, 3)
        strCycleName = sheet.cell_value(10, 7)
        ParallelProcess = sheet.cell_value(4, 7)
        strSequence = sheet.cell_value(4, 3)

        # jira Config sheets
        sheet = wb.sheet_by_name("JiraConfig")
        strJiraUrl = sheet.cell_value(1, 2)
        strJiraUserName = sheet.cell_value(2, 2)
        strJiraPassword = decrypt_data(sheet.cell_value(3, 2))
        # Delete custom log file if exist
        if os.path.exists(strProjectWorkingDirectory + "\CustomLogfile.txt"):
            os.remove(strProjectWorkingDirectory + "\CustomLogfile.txt")

        if os.path.exists(strProjectWorkingDirectory + "\PreReq_CustomLogfile.txt"):
            os.remove(strProjectWorkingDirectory + "\PreReq_CustomLogfile.txt")

        #Env Details
        sheet = wb.sheet_by_name("Env-Details")
        for row in range(sheet.nrows):
            if (sheet.cell(row,0).value == strEnvironment):
                strBrowser = str(sheet.cell(row, 1).value)
                strUrl = str(sheet.cell(row, 2).value)
                strDataSource = str(sheet.cell(row, 3).value)
                strDBUserName = str(sheet.cell(row, 4).value)
                strDBPassword = decrypt_data(str(sheet.cell(row, 5).value))
                ftp_server = str(sheet.cell(row, 6).value)
                ftp_username = str(sheet.cell(row, 7).value)
                ftp_password = decrypt_data(str(sheet.cell(row, 8).value))
                strRegionData = str(sheet.cell(row, 9).value)
                strEmailFrom = str(sheet.cell(row, 10).value)
                strEmailTo = str(sheet.cell(row,11).value)
                strEmailSubject = str(sheet.cell(row, 12).value)
                strAutosysUserName = str(sheet.cell(row, 13).value)
                strAutosysPassword = decrypt_data(str(sheet.cell(row, 14).value))
                strAutosysbaseUrl = str(sheet.cell(row, 15).value)
                break

        #All env Variables
        os.environ['strRunManagerFilePath']= strRunManagerFilePath
        os.environ['strRegion'] =strRegion
        os.environ['JiraProject'] =strJiraProject
        os.environ['VersionName'] =strVersionName
        os.environ['CycleName'] =strCycleName
        os.environ['JiraUrl'] =strJiraUrl
        os.environ['JiraUserName'] =strJiraUserName
        os.environ['JiraPassword'] = strJiraPassword
        os.environ['DataSheet'] = strModule + "_" + strRegion
        os.environ['ProjectDir'] = strProjectWorkingDirectory
        os.environ['Browser'] = strBrowser
        os.environ['Url'] = strUrl
        os.environ['DataSource'] = strDataSource
        os.environ['DBUser'] = strDBUserName
        os.environ['DBPwd'] = strDBPassword
        os.environ['Winscp_ServerName'] = ftp_server
        os.environ['Winscp_UserName'] = ftp_username
        os.environ['Winscp_Password'] = ftp_password
        os.environ['DataRegion'] = strRegionData
        os.environ['EmailFrom'] = strEmailFrom
        os.environ['EmailTo'] = strEmailTo
        os.environ['EmailSubject'] = strEmailSubject
        os.environ['AutosysUserName'] = strAutosysUserName
        os.environ['AutosysPassword'] = strAutosysPassword
        os.environ['AutosysbaseUrl'] = strAutosysbaseUrl
        os.environ['Process'] = str(int(ParallelProcess))

        # setting up Jira Configuration
        # Update the Jira No for each test case in Run Manager
        # df_config_Jira = pd.read_excel(strRunManagerFilePath, sheet_name=strRegion,usecols=['TestCase', 'Functionality', 'Start_Iteration', 'End_Iteration','Execute', 'ZephyrJira'])
        # df_config_Jira_Upd = JiraInitialSetup.Update_JiraList_RunManager(df_config_Jira)
        # with pd.ExcelWriter(strProjectWorkingDirectory+"\Config\RunManager.xlsx", engine='openpyxl') as writer:
        #     writer.book = load_workbook(strProjectWorkingDirectory+"\Config\RunManager.xlsx")
        #     df_config_Jira_Upd.to_excel(writer, sheet_name=strRegion, index=False)
        # wb = openpyxl.load_workbook(strProjectWorkingDirectory+"\Config\RunManager.xlsx")
        # ws1 = wb[strRegion+"1"]
        # ws2 = wb[strRegion]
        # for cell in ws1['F:F']:  # column F
        #     ws2.cell(row=cell.row, column=11, value=cell.value)
        # wb.remove(wb[strRegion+"1"])
        # wb.save(strProjectWorkingDirectory+"\Config\RunManager.xlsx")

        # Creating Test Cycle and Associating Zephyr Jira's
        df_config_Jira = pd.read_excel(strRunManagerFilePath, sheet_name=strRegion, usecols=['TestCase', 'Functionality', 'Start_Iteration', 'End_Iteration','Execute', 'ZephyrJira'])
        engine_Jira = create_engine('sqlite://', echo=False)
        df_config_Jira.to_sql('configJiraDB', engine_Jira, if_exists='replace', index=False)
        dbInfo_Func_Jira = """SELECT TestCase,ZephyrJira
                           FROM configJiraDB
                           WHERE Execute = 1 """
        df_Jira = engine_Jira.execute(dbInfo_Func_Jira)
        df_Jira = pd.DataFrame(list(df_Jira),columns=['TestCase', 'ZephyrJira'])
        lstZephyrJira = df_Jira['ZephyrJira'].values.tolist()
        if lstZephyrJira.count(None) == len(lstZephyrJira):
            print("Jira List is empty")
        else:
            strZephyrJira = ','.join(lstZephyrJira)
            print(strZephyrJira)
            os.environ['tests_to_add'] = strZephyrJira
            JiraInitialSetup.JiraSetup()



        #Generate_Notification_Email(strEmailSubject +" - Startup Notification", strEnvironment,date.today(),"Peoplesoft Test Execution Started. Please don't change dates in " + strEnvironment,"We will notify once execution completed.")

        # navigate to region sheet extract the Testcase name marked as True Functionality wise
        df_config = pd.read_excel(strRunManagerFilePath, sheet_name=strRegion,
                                  usecols=['TestCase', 'Functionality', 'Start_Iteration', 'End_Iteration',
                                           'Execute', 'ZephyrJira', 'SeqNum'])
        # df_config = df_config.query('Execute == True')
        engine = create_engine('sqlite://', echo=False)
        df_config.to_sql('configDB', engine, if_exists='replace', index=False)
        dbInfo_Functionality = """ SELECT Functionality
         FROM configDB
         WHERE Execute = 1 
         GROUP BY Functionality
         ORDER BY SeqNum ASC
         --ORDER BY count(Functionality) DESC"""
        DbInfo_Func_Result = engine.execute(dbInfo_Functionality)
        dfInfo_Func_Result = pd.DataFrame(list(DbInfo_Func_Result), columns=['Functionality'])
        dfInfo_Func_Result = dfInfo_Func_Result[dfInfo_Func_Result.Functionality != 'PreSanityCheck']
        dfInfo_Func_Result.loc[-1] = ['PreSanityCheck']
        dfInfo_Func_Result.index = dfInfo_Func_Result.index + 1  # shifting index
        dfInfo_Func_Result.sort_index(inplace=True)
        for func in dfInfo_Func_Result.itertuples(index=True, name='Pandas'):
            strFunctionalityName = func.Functionality
            print(str(strFunctionalityName))
            df_config_exec = pd.read_excel(strRunManagerFilePath, sheet_name=strRegion,
                                           usecols=['TestCase', 'Functionality', 'Start_Iteration', 'End_Iteration',
                                                    'Execute', 'ZephyrJira', 'SeqNum'])
            engine_exec = create_engine('sqlite://', echo=False)
            df_config_exec.to_sql('configExecDB', engine_exec, if_exists='replace', index=False)
            dbInfo_Func_Exec = """SELECT TestCase,Functionality,Start_Iteration,End_Iteration,Execute,ZephyrJira
                    FROM configExecDB
                    WHERE Functionality = '""" + strFunctionalityName + """'
                    AND Execute = 1 
                    ORDER BY SeqNum ASC"""
            df1 = engine_exec.execute(dbInfo_Func_Exec)
            df1 = pd.DataFrame(list(df1),
                               columns=['TestCase', 'Functionality', 'Start_Iteration', 'End_Iteration', 'Execute',
                                        'ZephyrJira'])
            lstTestCases = df1['TestCase'].values.tolist()
            strTestCases = ' -t '.join(lstTestCases)
            print(strFunctionalityName)


            #Create Result Folder
            strdatetime = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            strResultDirName = strProjectWorkingDirectory +'\Results\Results_' + strdatetime
            os.environ['ResultDirName'] = strResultDirName
            os.mkdir(strResultDirName)

            # Create Screenshot Folder
            strScreenShotDirName = strResultDirName + '\Screenshots'
            if os.path.isdir(strProjectWorkingDirectory+'\\Screenshots'):
                shutil.rmtree(strProjectWorkingDirectory+'\\Screenshots')
            os.mkdir(strScreenShotDirName)
            os.environ['ScreenShotDirName'] = strScreenShotDirName

            wbEmailLog = xlsxwriter.Workbook(strProjectWorkingDirectory + '\Results\Emaillog.xlsx')
            wsEmailLog = wbEmailLog.add_worksheet("EmailLog")
            os.environ['EmailLog'] = strProjectWorkingDirectory + '\Results\Emaillog.xlsx'
            wsEmailLog.write(0, 0, 'TestCase')
            wsEmailLog.write(0, 1, 'Iteration')
            wbEmailLog.close()

            # Execute Test Cases in Order
            if strSequence.casefold() == "YES".casefold():
                pabotSeqFilePath = os.path.join(strProjectWorkingDirectory, ".pabotsuitenames")
                if os.path.exists(pabotSeqFilePath):
                    os.remove(pabotSeqFilePath)
                with open(pabotSeqFilePath, "w+") as f:
                    f.write("datasources:320490b03c6b6afa4c6bc8c010d60582323017f5" + "\n")
                    f.write("commandlineoptions:eb6770a69da73c42cb30ae4a61eabc9a0ca98778" + "\n")
                    f.write("suitesfrom:no-suites-from-option" + "\n")
                    f.write("file:866f2157d49f18a19b908a1087a5c40b9f677be9" + "\n")
                    tcCount = 1
                    if len(lstTestCases) > 0:
                        for tc in lstTestCases:
                            if tcCount == 3:
                                f.write("#WAIT" + "\n")  # Need to put #WAIT statement as 2nd line in test case seq.
                            f.write("--test " + strFunctionalityName + "." + tc + "\n")
                            tcCount = tcCount + 1

            # create a command and run the robot file
            strCommand = "cmd.exe /c  \"cd " + strProjectWorkingDirectory + " & pabot --processes "+ str(int(ParallelProcess)) +" --testlevelsplit -t " + strTestCases + " -x junit.xml "+strProjectWorkingDirectory+"/TestSuite/" + strModule + "/" + strFunctionalityName + ".robot & robotmetrics --logo NomuraLogo.jpg -M metrics.html & python "+strProjectWorkingDirectory+"/Utilities/TestMetric.py"
            print(strCommand)
            os.system(strCommand)
            if strFunctionalityName == 'PreSanityCheck':
                with open(strProjectWorkingDirectory+"\\PreReq_CustomLogfile.txt") as file:
                    data = file.read()
                    if "FAIL" in data:
                        break

            #Generate_Notification_Email(strEmailSubject +" - End Notification", strEnvironment,date.today() ,"Peoplesoft Test Execution Completed in "+ strEnvironment,"Peoplesoft " +strEnvironment+" environment is now available for use.")



RunManager.ReadConfig()
# SendEmail()


